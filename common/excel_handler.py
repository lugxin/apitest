# !usr/bin/env python
# -*- coding:utf-8 _*-
"""
@version:
author:LongNight
@time: 2021/03/23
@file: excel_handler.py
@function:
@modify:
"""
from openpyxl import load_workbook

# import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler():

    def __init__(self, file):
        self.file = file

    def open_sheet(self, name) -> Worksheet:
        # 打开表单
        wb = load_workbook(self.file)
        sheet = wb[name]
        wb.close()
        return sheet

    def header(self, sheet_name):
        sheet = self.open_sheet(sheet_name)
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    def read(self, sheet_name):
        # 读取所有的数据
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows)
        # 获取标题

        data = []
        for row in rows[1:]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
                # 列表转成字典：要和header 去zip
                data_dict = dict(zip(self.header(sheet_name), row_data))
            data.append(data_dict)
        return data

    # @staticmethod
    def write(self, file, sheet_name, row, column, data):
        # 写入Excel 数据
        wb = load_workbook(file)
        sheet = wb[sheet_name]
        # 修改单元格
        sheet.cell(row, column).value = data

        # 保存
        wb.save(file)
        # 关闭
        wb.close()


if __name__ == '__main__':
    excel = ExcelHandler(r'd:\demo.xlsx')
    excel.write(r'd:\deme.xlsx', 'Sheet', 3, 1, 'amoshows')
    data = excel.read('Sheet1')
    print(data)
